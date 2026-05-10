import pandas as pd
import re
import logging
import time
import random
import smtplib
import requests
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from openpyxl import load_workbook
from openpyxl.styles import Font
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import os
from pathlib import Path

# ================= LOGGING SETUP =================
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s | %(levelname)-7s | %(message)s')
file_handler = logging.FileHandler('imot_scraper.log', encoding='utf-8')
file_handler.setLevel(logging.INFO)
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

logger.info("=== Starting imot.bg scraper ===")

# ================= PATHS & CONFIG =================
HISTORY_FILE = "all_listings_history.parquet"
HTML_OUTPUT = Path("docs/index.html")  # GitHub Pages serves from /docs
IMAGES_DIR = Path("docs/images")  # Downloaded listing images
HTML_OUTPUT.parent.mkdir(exist_ok=True)
IMAGES_DIR.mkdir(parents=True, exist_ok=True)

TODAY = datetime.now().strftime("%Y-%m-%d")
NOW_STR = datetime.now().strftime("%d.%m.%Y %H:%M")
timestamp = datetime.now().strftime("%Y%m%d_%H%M")
excel_file = f"imot_bg_scraping_{timestamp}.xlsx"

# ── Email ─────────────────────────────────────────────────────────────────────
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SENDER_EMAIL = os.environ.get("SENDER_EMAIL")
SENDER_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD")
if not SENDER_PASSWORD:
    logger.error("GMAIL_APP_PASSWORD secret is not set!")

# Списък с получатели от променлива на средата (разделени със запетая)
RECEIVERS_RAW = os.environ.get("RECEIVERS")
RECEIVERS = [r.strip() for r in RECEIVERS_RAW.split(",") if r.strip()]

# ── Search URL ────────────────────────────────────────────────────────────────
# Линкът за търсене вече се взима от променлива на средата
base_url = os.environ.get(
    "BASE_URL",
)

listings = []

# ── Column constants ──────────────────────────────────────────────────────────
COL_LINK = 'Link'
COL_PRICE = 'Price_EUR'
COL_SIZE = 'Size_sqm'
COL_PRICE_PER_SQM = 'Price_EUR_per_sqm'
COL_LOCATION = 'Location'
COL_FLOOR = 'Floor'
COL_TOTAL_FLOORS = 'Total_floors'
COL_YEAR = 'Year_built'
COL_INFO = 'Info'
COL_TITLE = 'Title'
COL_SCRAPED_DATE = 'Scraped_Date'
COL_FIRST_SEEN = 'First_Seen_Date'
COL_PRICE_HISTORY = 'Price_History'
COL_SOLD = 'Sold'
COL_SITE_PRICE_HISTORY = 'Свалена_ценова_история'
COL_IMAGES = 'Image_Paths'  # comma-separated relative paths


# ================= HELPERS =================

def clean_price(raw):
    if not raw: return None
    m = re.search(r'([\d\s.,]+)\s*€', raw)
    if not m: return None
    cleaned = m.group(1).replace(' ', '').replace(',', '.')
    try:
        return float(cleaned)
    except ValueError:
        logger.warning(f"Failed to parse price: {raw}")
        return None


def normalize_location(loc):
    if not loc: return loc
    rules = {
        r'Младост\s*IV-ти': 'Младост 4',
        r'Младост\s*IV': 'Младост 4',
        r'Младост\s*V': 'Младост 5',
        r'Младост\s*III': 'Младост 3',
        r'Младост\s*II': 'Младост 2',
        r'Младост\s*I\b': 'Младост 1',
    }
    orig = loc
    for pat, repl in rules.items():
        loc = re.sub(pat, repl, loc, flags=re.IGNORECASE)
    if orig != loc:
        logger.debug(f"Normalized location: {orig} → {loc}")
    return loc


def parse_total_ads(page_content):
    soup = BeautifulSoup(page_content, 'html.parser')
    text = soup.get_text(separator=' ', strip=True)
    m = re.search(r'(?:от\s*общо\s*|от\s*)(\d+)\s*обяв[иа]', text, re.IGNORECASE)
    if m: return int(m.group(1))
    m2 = re.search(r'(\d+)\s*[-–]\s*\d+\s*от\s*общо\s*(\d+)', text)
    if m2: return int(m2.group(2))
    return None


def parse_page(page_content, pg_num=None):
    soup = BeautifulSoup(page_content, 'html.parser')
    items = soup.find_all('div', class_='item')
    if not items:
        logger.warning(f"No items found on page {pg_num or 1}")
        return 0

    page_count = 0
    for item in items:
        try:
            title_a = item.find('a', class_='title')
            if not title_a: continue

            title = title_a.get_text(separator=' ', strip=True)
            href = title_a['href'].strip()
            if 'fakti.bg' in href.lower() or not href.startswith(('https://www.imot.bg', '//', '/')):
                continue
            if href.startswith('//'):
                href = 'https:' + href
            elif href.startswith('/'):
                href = 'https://www.imot.bg' + href
            elif not href.startswith('http'):
                href = 'https://www.imot.bg/' + href

            price_div = item.find('div', class_='price')
            price_raw = price_div.get_text(strip=True) if price_div else ''
            price_eur = clean_price(price_raw)

            info_div = item.find('div', class_='info')
            info_text = info_div.get_text(strip=True) if info_div else ''

            size = None
            m = re.search(r'(\d{2,3})\s*кв\.?\s*м', info_text)
            if m: size = int(m.group(1))

            floor = total_floors = None
            m_floor = re.search(r'(\d+)\s*[-–]\s*(?:ти|ри)?', info_text)
            if m_floor: floor = int(m_floor.group(1))
            m_total = re.search(r'(?:от\s*|/\s*)(\d+)', info_text)
            if m_total: total_floors = int(m_total.group(1))

            year = None
            m_year = re.search(r'(19\d{2}|20\d{2})(?:\s*[-–]\s*(19\d{2}|20\d{2}))?', info_text)
            if m_year: year = m_year.group(0).strip()

            location = None
            if 'град София,' in title:
                location = title.split('град София,')[-1].strip()
                location = normalize_location(location)

            price_m2 = round(price_eur / size, 2) if price_eur and size and size > 0 else None

            listings.append({
                COL_TITLE: title,
                COL_LOCATION: location,
                COL_PRICE: price_eur,
                COL_SIZE: size,
                COL_PRICE_PER_SQM: price_m2,
                COL_FLOOR: floor,
                COL_TOTAL_FLOORS: total_floors,
                COL_YEAR: year,
                COL_INFO: info_text,
                COL_LINK: href,
                COL_SCRAPED_DATE: TODAY,
                COL_FIRST_SEEN: TODAY,
                COL_PRICE_HISTORY: "",
                COL_SITE_PRICE_HISTORY: "",
                COL_IMAGES: "",
            })
            page_count += 1
        except Exception as parse_err:
            logger.warning(f"Error parsing item on page {pg_num or 1}: {parse_err}")

    logger.info(f"Scraped {page_count} listings from page {pg_num or 1}")
    return page_count


def format_price_history_entry(price, date_str):
    if pd.isna(price) or price is None or pd.isna(date_str):
        return ""
    try:
        price_int = int(round(float(price)))
        return f"{price_int:,} € ({date_str})"
    except (ValueError, TypeError):
        return ""


def append_current_if_needed(hist, price, date):
    if pd.isna(price) or pd.isna(date):
        return hist if isinstance(hist, str) else ""
    curr = format_price_history_entry(price, date)
    if not isinstance(hist, str):
        hist = ""
    if curr and curr not in hist:
        return f"{hist} → {curr}" if hist.strip() else curr
    return hist


def deduplicate_history(df, link_col, price_history_col):
    def merge_history(series):
        vals = [v.strip() for v in series if isinstance(v, str) and v.strip()]
        seen = set()
        result = []
        for v in vals:
            if v not in seen:
                seen.add(v)
                result.append(v)
        return " → ".join(result) if result else ""

    df[price_history_col] = df[price_history_col].fillna("").astype(str)
    agg = {}
    for col_key in df.columns:
        agg[col_key] = merge_history if col_key == price_history_col else 'last'
    return df.groupby(link_col, as_index=False).agg(agg)


def parse_site_price_history_html(raw_html):
    if not raw_html or len(raw_html) < 100:
        return ""

    soup = BeautifulSoup(raw_html, 'html.parser')

    # Търсим контейнера по различни начини
    container = soup.find('div', id='priceHistory2')
    if not container:
        return ""

    statistiki = container.find('statistiki')
    if not statistiki:
        statistiki = soup.find('statistiki')
    if not statistiki:
        return ""

    divs = statistiki.find_all('div', recursive=False)
    if len(divs) < 10:
        logger.debug(f"Недостатъчно div-ове в price history: {len(divs)}")
        return ""

    def clean(el):
        if not el:
            return ""
        return el.get_text(separator=" ", strip=True).replace("\xa0", " ").strip()

    parts = []
    # Пропускаме header-а (обикновено първите 4 div-а)
    data_divs = divs[4:]

    for i in range(0, len(data_divs), 4):
        if i + 2 >= len(data_divs):
            break
        date_txt = clean(data_divs[i])
        change_txt = clean(data_divs[i + 1])
        price_txt = clean(data_divs[i + 2])

        if not price_txt or "€" not in price_txt:
            continue

        if any(x in date_txt.lower() for x in ["начало", "начална"]):
            parts.append(f"Начална: {price_txt}")
        else:
            span_class = ""
            if "-" in change_txt:
                span_class = ' class="price-down"'
            elif "+" in change_txt:
                span_class = ' class="price-up"'

            change_part = f' <span{span_class}>{change_txt}</span>' if change_txt and change_txt not in ["", "—"] else ""
            parts.append(f"{date_txt}{change_part} → {price_txt}")

    result = " | ".join(parts)
    if result:
        logger.info(f"Успешно извлечена история: {result[:120]}...")
    else:
        logger.warning("Не успях да извлека ценова история от страницата")

    return result


# ================= IMAGE DOWNLOAD =================

def get_listing_id_from_url(url):
    """Извлича ID-то на обявата от URL-а (напр. 1c176432069391545)"""
    match = re.search(r'obiava-([a-z0-9]+)-', url)
    return match.group(1) if match else None


def extract_images_improved(page, listing_url, max_images=2):
    urls = []
    listing_id = get_listing_id_from_url(listing_url)

    if not listing_id:
        logger.warning("Could not extract listing ID from URL")
        return []

    try:
        page.wait_for_timeout(1500)

        for _ in range(5):
            page.mouse.wheel(0, 1000)
            page.wait_for_timeout(500)

        # Взимаме всички carousel изображения
        images = page.query_selector_all("img.carouselimg")

        # Събираме кандидатите с техния номер от alt="... изображение N"
        candidates_with_id = []
        candidates_fallback = []
        for img in images:
            try:
                data_src_gallery = img.get_attribute("data-src-gallery")
                data_src = img.get_attribute("data-src")
                src = img.get_attribute("src")
                alt = img.get_attribute("alt") or ""

                candidate = data_src_gallery or data_src or src

                if candidate and ("imotstatic" in candidate or "cdn" in candidate or "focus.bg" in candidate):
                    m = re.search(r'изображение\s+(\d+)', alt, re.IGNORECASE)
                    order = int(m.group(1)) if m else 9999

                    if listing_id in candidate:
                        candidates_with_id.append((order, candidate))
                    else:
                        candidates_fallback.append((order, candidate))
            except:
                continue

        # Ако точните съвпадения има – ползваме тях, иначе – fallback без ID
        candidates = candidates_with_id if candidates_with_id else candidates_fallback
        if candidates_fallback and not candidates_with_id:
            logger.info(f"  ℹ listing_id not in CDN URLs, using fallback images for {listing_id}")

        # Сортираме по номер → изображение 1 първо, изображение 2 второ и т.н.
        candidates.sort(key=lambda x: x[0])
        urls = [url for _, url in candidates]

        # Премахваме дубликати и лимитираме
        unique_urls = list(dict.fromkeys(urls))
        logger.debug(
            f"Found {len(unique_urls)} images for listing {listing_id}, "
            f"order: {[o for o, _ in candidates[:max_images]]}"
        )

        return unique_urls[:max_images]

    except Exception as e:
        logger.warning(f"extract_images_improved error: {e}")
        return urls


def download_images_from_urls(listing_url, image_urls, max_images=2):
    if not image_urls:
        logger.warning(f"No valid images found for {listing_url}")
        return ""

    lid = get_listing_id_from_url(listing_url)
    img_dir = IMAGES_DIR / lid
    img_dir.mkdir(parents=True, exist_ok=True)

    saved = []
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

    for i, src in enumerate(image_urls[:max_images], start=1):
        try:
            r = requests.get(src, timeout=20, headers=headers)
            if r.status_code == 200 and len(r.content) > 10000:  # филтър за качествени снимки
                path = img_dir / f"{i}.jpg"
                path.write_bytes(r.content)
                rel = f"images/{lid}/{i}.jpg"
                saved.append(rel)
                logger.debug(f"✓ Downloaded correct image {i} for {listing_url}")
        except Exception as ex:
            logger.debug(f"Download failed: {ex}")

    return ",".join(saved) if saved else ""


# ================= SELENIUM + PLAYWRIGHT: PRICE HISTORY + IMAGES =================

def scrape_site_price_histories_selenium(links):
    if not links:
        return {}

    total = len(links)
    result = {url: {"price_history": "", "images": ""} for url in links}

    logger.info(f"Processing {total} listings...")

    # Selenium за price history
    driver = None
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC

        options = Options()
        options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")

        driver = webdriver.Chrome(options=options)
        wait = WebDriverWait(driver, 10)
    except Exception as selenium_err:
        logger.error(f"Selenium failed: {selenium_err}")

    # Playwright за снимки
    pw_browser = None
    try:
        from playwright.sync_api import sync_playwright
        pw = sync_playwright().start()
        pw_browser = pw.chromium.launch(headless=True)
    except Exception as pw_init_err:
        logger.warning(f"Playwright init failed: {pw_init_err}")

    good_listings = 0

    for idx, url in enumerate(links, start=1):
        logger.info(f"[{idx}/{total}] {url}")

        # ================= PRICE HISTORY =================
        price_hist = ""

        if driver:
            try:
                driver.get(url)

                # Изчакваме контейнера
                price_history_elem = wait.until(
                    EC.presence_of_element_located((By.ID, "priceHistory2"))
                )

                # зарежда history-то чрез JS
                try:
                    title_span = price_history_elem.find_element(
                        By.CSS_SELECTOR,
                        'div.title span[onclick*="showpricechange"]',
                    )

                    onclick_attr = title_span.get_attribute("onclick") or ""

                    start = onclick_attr.find("(") + 1
                    end = onclick_attr.rfind(");")

                    params_str = onclick_attr[start:end]
                    params = [p.strip().strip("'") for p in params_str.split(",")]

                    if len(params) >= 5:
                        js_code = (
                            f"showpricechange('{params[0]}','{params[1]}',"
                            f"'{params[2]}','{params[3]}','{params[4]}');"
                        )

                        driver.execute_script(js_code)

                        try:
                            wait.until(
                                EC.presence_of_element_located(
                                    (By.TAG_NAME, "statistiki")
                                )
                            )
                        except Exception as price_hist_err:
                            logger.warning(f"... {price_hist_err}")

                        time.sleep(1)

                except Exception as js_err:
                    logger.debug(f"showpricechange failed: {js_err}")

                page_html = driver.page_source
                price_hist = parse_site_price_history_html(page_html)

                if price_hist:
                    logger.info(f"  ✔ История на цената: {price_hist[:150]}...")
                    result[url]["price_history"] = price_hist
                else:
                    logger.warning(f"  ⚠ Няма извлечена история за {url}")
                    # Запази поне празен string
                    result[url]["price_history"] = ""


            except Exception as hist_err:
                logger.warning(f"Price history error for {url}: {hist_err}")
                # Ако е Chrome crash (празен Message:), рестартираме driver-а
                if "Message:" in str(hist_err) and len(str(hist_err).strip()) < 20:
                    logger.warning("  Chrome crash detected – restarting Selenium driver")
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    try:
                        driver = webdriver.Chrome(options=options)
                        wait = WebDriverWait(driver, 10)
                        logger.info("  ✔ Selenium driver restarted")
                    except Exception as restart_err:
                        logger.error(f"  Could not restart driver: {restart_err}")
                        driver = None

        # ================= IMAGES (Playwright) =================
        img_paths = ""
        if pw_browser:
            try:
                p_page = pw_browser.new_page()
                p_page.set_extra_http_headers({
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
                })
                p_page.goto(url, wait_until="domcontentloaded", timeout=30000)
                image_urls = extract_images_improved(p_page, url, max_images=2)
                if image_urls:
                    img_paths = download_images_from_urls(url, image_urls, max_images=2)
                    if img_paths:
                        logger.info(f"  → {len(img_paths.split(','))} снимки свалени")
                    else:
                        logger.warning(f"  → Не успях да сваля снимките")
                else:
                    logger.warning(f"  → Няма намерени снимки за обявата")

                p_page.close()

            except Exception as pw_err:
                logger.warning(f"Playwright error for {url}: {pw_err}")

        result[url]["images"] = img_paths

        time.sleep(2 + random.uniform(0, 1))

    # Cleanup
    if driver:
        driver.quit()
    if pw_browser:
        pw_browser.close()

    logger.info(f"Finished. Good listings with images: {good_listings}/{total}")
    return result


# ================= HTML GENERATOR =================

def _fmt_price(x):
    return f"{int(round(x)):,} €".replace(",", "\u202f") if pd.notna(x) and x else "—"


def _fmt_size(x):
    return f"{int(x)} m²" if pd.notna(x) and x else "—"


def _fmt_pm2(x):
    return f"{int(round(x))} €/m²" if pd.notna(x) and x else "—"


def _link_cell(href):
    if not href or pd.isna(href): return "—"
    return f'<a href="{href}" target="_blank" rel="noopener">🔗 Виж</a>'


def _img_cell(paths_str):
    """Показва до 2 thumbnail-а от запазените снимки."""
    if not paths_str or pd.isna(paths_str) or str(paths_str).strip() == "":
        return "—"
    html_parts = []
    for img_path in str(paths_str).split(",")[:2]:
        img_path = img_path.strip()
        if img_path:
            html_parts.append(
                f'<img src="{img_path}" '
                f'style="width:72px;height:54px;object-fit:cover;'
                f'border-radius:4px;margin-right:4px;cursor:pointer" '
                f'onclick="window.open(this.src)" '
                f'onerror="this.style.display=\'none\'">'
            )
    return "".join(html_parts) if html_parts else "—"


def _build_rows(df, cols):
    rows_html = []
    for _, row in df.iterrows():
        cells = []
        for col_key in cols:
            val = row.get(col_key, "")

            if col_key == COL_LINK:
                cells.append(f"<td>{_link_cell(val)}</td>")

            elif col_key == COL_PRICE:
                cells.append(f"<td>{_fmt_price(val)}</td>")

            elif col_key == COL_SIZE:
                cells.append(f"<td>{_fmt_size(val)}</td>")

            elif col_key == COL_PRICE_PER_SQM:
                cells.append(f"<td>{_fmt_pm2(val)}</td>")

            elif col_key == COL_IMAGES:
                cells.append(f"<td>{_img_cell(val)}</td>")

            elif col_key == COL_SITE_PRICE_HISTORY:
                text = str(val).strip() if pd.notna(val) else ""
                if text and text != "—":
                    cells.append(f'<td><div class="history site-history">{text}</div></td>')
                else:
                    cells.append('<td>—</td>')

            elif col_key == COL_PRICE_HISTORY:
                text = str(val).strip() if pd.notna(val) else ""
                if text and text != "—":
                    cells.append(f'<td><span class="history">{text}</span></td>')
                else:
                    cells.append('<td>—</td>')

            else:
                cells.append(f"<td>{val if pd.notna(val) and val != '' else '—'}</td>")

        rows_html.append("<tr>" + "".join(cells) + "</tr>")
    return "\n".join(rows_html)


def _table(df, cols, headers, css_id="", extra_class=""):
    if df.empty:
        return '<p class="empty-note">Няма данни.</p>'
    thead = "<tr>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr>"
    tbody = _build_rows(df, cols)
    id_attr = f' id="{css_id}"' if css_id else ""
    cls = f'data-table {extra_class}'.strip()
    return f'<table{id_attr} class="{cls}"><thead>{thead}</thead><tbody>{tbody}</tbody></table>'


def generate_html(df_input: pd.DataFrame, now_str: str):

    # ── Derive sections ───────────────────────────────────────────────────────
    df_active = df_input[~df_input[COL_SOLD].fillna(False)].copy() if not df_input.empty else pd.DataFrame()
    df_new_all = df_active.copy()

    if not df_active.empty and COL_PRICE_HISTORY in df_active.columns:
        mask_changed = df_active[COL_PRICE_HISTORY].fillna("").str.contains(" → ")
        df_changed_all = df_active[mask_changed].copy()
    else:
        df_changed_all = pd.DataFrame()

    df_sold = df_input[df_input[COL_SOLD].fillna(False)].copy() if not df_input.empty else pd.DataFrame()

    n_total = len(df_active)
    n_changed = len(df_changed_all)
    n_sold = len(df_sold)

    if not df_active.empty:
        df_active = df_active.sort_values(COL_SCRAPED_DATE, ascending=False)
    if not df_new_all.empty:
        df_new_all = df_new_all.sort_values(COL_FIRST_SEEN if COL_FIRST_SEEN in df_new_all.columns else COL_SCRAPED_DATE,
                                        ascending=False)
    if not df_changed_all.empty:
        df_changed_all = df_changed_all.sort_values(COL_SCRAPED_DATE, ascending=False)

    # ── Build tables ─────────────────────────────────────────────────────
    new_table = _table(
        df_new_all,
        [COL_IMAGES, COL_LOCATION, COL_PRICE, COL_SIZE, COL_PRICE_PER_SQM,
         COL_FLOOR, COL_TOTAL_FLOORS, COL_YEAR, COL_SITE_PRICE_HISTORY, COL_LINK],
        ["Снимки", "Локация", "Цена", "Площ", "€/m²", "Ет.", "Общо ет.", "Година",
         "Свалена ценова история", ""],
        css_id="new-table",
    )

    changed_table = _table(
        df_changed_all,
        [COL_IMAGES, COL_LOCATION, COL_PRICE, COL_SIZE, COL_PRICE_PER_SQM,
         COL_FLOOR, COL_TOTAL_FLOORS, COL_YEAR, COL_SCRAPED_DATE,
         COL_PRICE_HISTORY, COL_SITE_PRICE_HISTORY, COL_LINK],
        ["Снимки", "Локация", "Цена", "Площ", "€/m²", "Ет.", "Общо ет.", "Год.",
         "Последно виждана", "История на цената", "Свалена ценова история", ""],
        css_id="changed-table",
    )

    all_table = _table(
        df_active,
        [COL_IMAGES, COL_LOCATION, COL_PRICE, COL_SIZE, COL_PRICE_PER_SQM,
         COL_FLOOR, COL_TOTAL_FLOORS, COL_YEAR, COL_SCRAPED_DATE,
         COL_PRICE_HISTORY, COL_SITE_PRICE_HISTORY, COL_LINK],
        ["Снимки", "Локация", "Цена", "Площ", "€/m²", "Ет.", "Общо ет.", "Год.",
         "Последно виждана", "История на цената", "Свалена ценова история", ""],
        css_id="all-table",
    )

    sold_table = _table(
        df_sold,
        [COL_IMAGES, COL_LOCATION, COL_PRICE, COL_SIZE, COL_PRICE_PER_SQM,
         COL_SCRAPED_DATE, COL_PRICE_HISTORY, COL_SITE_PRICE_HISTORY, COL_LINK],
        ["Снимки", "Локация", "Последна цена", "Площ", "€/m²",
         "Последно виждана", "История на цената", "Свалена ценова история", ""],
        css_id="sold-table",
        extra_class="sold-table",
    )

    # ── Full HTML ─────────────────────────────────────────────────────────────
    dashboard_html = f"""<!DOCTYPE html>
<html lang="bg">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Имоти сем. Кирилови</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap" rel="stylesheet">
<style>
  :root {{
    --bg:        #0f1117;
    --surface:   #1a1d27;
    --border:    #2a2d3a;
    --text:      #e2e4f0;
    --muted:     #7a7d9a;
    --accent:    #4f9cf9;
    --green:     #3ecf8e;
    --orange:    #f59e0b;
    --red:       #f43f5e;
    --mono:      'IBM Plex Mono', monospace;
    --sans:      'IBM Plex Sans', sans-serif;
  }}

  * {{ box-sizing: border-box; margin: 0; padding: 0; }}

  body {{
    font-family: var(--sans);
    background: var(--bg);
    color: var(--text);
    font-size: 14px;
    line-height: 1.6;
  }}

  /* ── Header ── */
  header {{
    padding: 28px 32px 20px;
    border-bottom: 1px solid var(--border);
    display: flex;
    align-items: baseline;
    gap: 24px;
    flex-wrap: wrap;
  }}
  header h1 {{
    font-size: 22px;
    font-weight: 600;
    letter-spacing: -0.5px;
    color: var(--text);
  }}
  header h1 span {{ color: var(--accent); }}
  .updated {{
    font-family: var(--mono);
    font-size: 12px;
    color: var(--muted);
    margin-left: auto;
  }}

  /* ── Stats bar ── */
  .stats {{
    display: flex;
    gap: 1px;
    background: var(--border);
    border-bottom: 1px solid var(--border);
  }}
  .stat {{
    flex: 1;
    padding: 18px 24px;
    background: var(--surface);
    text-align: center;
  }}
  .stat-num {{
    font-family: var(--mono);
    font-size: 28px;
    font-weight: 600;
    line-height: 1;
    color: var(--accent);
  }}
  .stat-num.green  {{ color: var(--green); }}
  .stat-num.orange {{ color: var(--orange); }}
  .stat-num.red    {{ color: var(--red); }}
  .stat-label {{
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    color: var(--muted);
    margin-top: 4px;
  }}

  /* ── Nav tabs ── */
  nav {{
    padding: 0 32px;
    border-bottom: 1px solid var(--border);
    display: flex;
    gap: 0;
    overflow-x: auto;
  }}
  nav a {{
    display: inline-block;
    padding: 14px 20px;
    font-size: 13px;
    font-weight: 600;
    color: var(--muted);
    text-decoration: none;
    border-bottom: 2px solid transparent;
    white-space: nowrap;
    transition: color 0.15s, border-color 0.15s;
  }}
  nav a:hover  {{ color: var(--text); border-color: var(--border); }}
  nav a.active {{ color: var(--accent); border-color: var(--accent); }}

  /* ── Sections ── */
  main {{ padding: 0 32px 48px; }}
  section {{ padding-top: 36px; }}
  section h2 {{
    font-size: 15px;
    font-weight: 600;
    color: var(--text);
    margin-bottom: 4px;
    display: flex;
    align-items: center;
    gap: 10px;
  }}
  section h2 .badge {{
    font-family: var(--mono);
    font-size: 11px;
    padding: 1px 8px;
    border-radius: 99px;
    background: var(--border);
    color: var(--muted);
  }}
  .section-desc {{
    font-size: 12px;
    color: var(--muted);
    margin-bottom: 16px;
  }}

  /* ── Search bar ── */
  .search-wrap {{
    margin-bottom: 12px;
  }}
  .search-wrap input {{
    width: 320px;
    max-width: 100%;
    padding: 8px 14px;
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 6px;
    color: var(--text);
    font-family: var(--sans);
    font-size: 13px;
    outline: none;
    transition: border-color 0.15s;
  }}
  .search-wrap input:focus {{ border-color: var(--accent); }}
  .search-wrap input::placeholder {{ color: var(--muted); }}

  /* ── Tables ── */
  .table-wrap {{
    overflow-x: auto;
    border: 1px solid var(--border);
    border-radius: 8px;
  }}
  table.data-table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 13px;
  }}
  table.data-table thead {{
    position: sticky;
    top: 0;
    z-index: 1;
  }}
  table.data-table th {{
    background: var(--surface);
    color: var(--muted);
    font-weight: 600;
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    padding: 10px 14px;
    border-bottom: 1px solid var(--border);
    white-space: nowrap;
    cursor: pointer;
    user-select: none;
  }}
  table.data-table th:hover {{ color: var(--text); }}
  table.data-table th.sorted-asc::after  {{ content: ' ↑'; }}
  table.data-table th.sorted-desc::after {{ content: ' ↓'; }}

  table.data-table td {{
    padding: 10px 14px;
    border-bottom: 1px solid var(--border);
    vertical-align: middle;
    max-width: 280px;
    word-break: break-word;
  }}
  table.data-table tr:last-child td {{ border-bottom: none; }}
  table.data-table tr:hover td {{ background: #1e2130; }}

  table.data-table td.old-price {{
    color: var(--muted);
    text-decoration: line-through;
  }}

  table.sold-table td {{ color: var(--muted); }}
  table.sold-table td a {{ color: var(--muted); }}
  table.sold-table img {{ opacity: 0.5; }}

  table.data-table a {{
    color: var(--accent);
    text-decoration: none;
    font-weight: 600;
  }}
  table.data-table a:hover {{ text-decoration: underline; }}

  /* price history cell — smaller mono */
  table.data-table td:has(.history) {{ font-size: 11px; }}
.history {{
    font-family: var(--mono);
    font-size: 11px;
    line-height: 1.35;
    color: #a0a3c0;
  }}
  .site-history {{
    color: #7dd3fc;
  }}
  .price-down {{ color: var(--green) !important; font-weight: 600; }}
  .price-up {{ color: var(--red) !important; font-weight: 600; }}

  /* thumbnail images */
  table.data-table td img {{
    border: 1px solid var(--border);
    border-radius: 4px;
    transition: transform 0.15s;
  }}
  table.data-table td img:hover {{
    transform: scale(1.06);
  }}

  .empty-note {{
    color: var(--muted);
    font-size: 13px;
    padding: 20px 0;
  }}

  /* ── Footer ── */
  footer {{
    text-align: center;
    padding: 24px;
    font-size: 12px;
    color: var(--muted);
    border-top: 1px solid var(--border);
  }}

  @media (max-width: 640px) {{
    header {{ padding: 16px; }}
    main   {{ padding: 0 16px 32px; }}
    nav    {{ padding: 0 16px; }}
    .stats {{ flex-wrap: wrap; }}
    .stat  {{ flex: 1 1 50%; }}
    .stat-num {{ font-size: 22px; }}
  }}
</style>
</head>
<body>

<header>
  <h1>🏠 Имоти · <span>сем. Кирилови</span></h1>
  <span class="updated">Обновено: {now_str}</span>
</header>

<div class="stats">
  <div class="stat">
    <div class="stat-num">{n_total}</div>
    <div class="stat-label">Активни обяви</div>
  </div>
  <div class="stat">
    <div class="stat-num orange">{n_changed}</div>
    <div class="stat-label">Промени в цена</div>
  </div>
  <div class="stat">
    <div class="stat-num red">{n_sold}</div>
    <div class="stat-label">Продадени / свалени</div>
  </div>
</div>

<nav>
  <a href="#all"     class="active">Всички активни</a>
  <a href="#changed">Промени</a>
  <a href="#sold">Продадени</a>
</nav>

<main>

  <section id="all">
    <h2>Всички активни обяви <span class="badge">{n_total}</span></h2>
    <p class="section-desc">Пълен списък на текущо активните обяви.</p>
    <div class="search-wrap">
      <input type="text" id="all-search" placeholder="Търси по локация, цена, площ…" oninput="filterTable('all-table', this.value)">
    </div>
    <div class="table-wrap">
      {all_table}
    </div>
  </section>

  <section id="changed">
    <h2>Промени в цената <span class="badge">{n_changed}</span></h2>
    <p class="section-desc">Обяви с регистрирана промяна в цената.</p>
    <div class="table-wrap">
      {changed_table}
    </div>
  </section>

  <section id="sold">
    <h2>Продадени / свалени <span class="badge">{n_sold}</span></h2>
    <p class="section-desc">Обяви, изчезнали от сайта (вероятно продадени или свалени).</p>
    <div class="table-wrap">
      {sold_table}
    </div>
  </section>

</main>

<footer>
  Данни от imot.bg · Обновява се автоматично в 07:00 и 13:00 ч. · {now_str}
</footer>

<script>
// ── Simple table search ──────────────────────────────────────────────────────
function filterTable(tableId, query) {{
  const tbl  = document.getElementById(tableId);
  if (!tbl) return;
  const rows = tbl.querySelectorAll('tbody tr');
  const q    = query.trim().toLowerCase();
  rows.forEach(r => {{
    r.style.display = q === '' || r.textContent.toLowerCase().includes(q) ? '' : 'none';
  }});
}}

// ── Sortable columns ─────────────────────────────────────────────────────────
document.querySelectorAll('table.data-table th').forEach((th, colIdx) => {{
  th.addEventListener('click', () => {{
    const table = th.closest('table');
    const tbody = table.querySelector('tbody');
    const rows  = Array.from(tbody.querySelectorAll('tr'));
    const asc   = th.classList.contains('sorted-asc');

    table.querySelectorAll('th').forEach(t => t.classList.remove('sorted-asc','sorted-desc'));
    th.classList.add(asc ? 'sorted-desc' : 'sorted-asc');

    rows.sort((a, b) => {{
      const va = a.children[colIdx]?.textContent.trim() || '';
      const vb = b.children[colIdx]?.textContent.trim() || '';
      const na = parseFloat(va.replace(/[^\\d.]/g,''));
      const nb = parseFloat(vb.replace(/[^\\d.]/g,''));
      const cmp = !isNaN(na) && !isNaN(nb)
        ? na - nb
        : va.localeCompare(vb, 'bg');
      return asc ? -cmp : cmp;
    }});
    rows.forEach(r => tbody.appendChild(r));
  }});
}});

// ── Nav highlight on scroll ──────────────────────────────────────────────────
const sections = document.querySelectorAll('main section[id]');
const navLinks = document.querySelectorAll('nav a');
const observer = new IntersectionObserver(entries => {{
  entries.forEach(e => {{
    if (e.isIntersecting) {{
      navLinks.forEach(a => a.classList.remove('active'));
      const link = document.querySelector('nav a[href="#' + e.target.id + '"]');
      if (link) link.classList.add('active');
    }}
  }});
}}, {{ threshold: 0.25 }});
sections.forEach(s => observer.observe(s));
</script>

</body>
</html>"""

    HTML_OUTPUT.write_text(dashboard_html, encoding="utf-8")
    logger.info(f"HTML dashboard written to {HTML_OUTPUT}")


# ================= SCRAPING =================
with sync_playwright() as p:
    logger.info("Opening browser (headless)…")
    browser = p.chromium.launch(
        headless=True,
        args=["--no-sandbox", "--disable-dev-shm-usage"]
    )
    page = browser.new_page()
    page.set_extra_http_headers({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    })
    total_ads = None
    items_per_page = 40

    try:
        logger.info(f"Loading first page: {base_url}")
        page.goto(base_url, wait_until='domcontentloaded', timeout=60000)
        page.wait_for_selector('div.item', timeout=25000)

        html_content = page.content()
        total_ads = parse_total_ads(html_content)
        if total_ads:
            logger.info(f"Total listings: {total_ads}")
        else:
            logger.warning("Could not extract total listings count")

        parse_page(html_content, pg_num=1)

        max_pages = (
            (total_ads // items_per_page) + (1 if total_ads % items_per_page else 0)
            if total_ads else 5
        )
        logger.info(f"Planning to scrape up to {max_pages} pages")

        pg = 2
        while pg <= max_pages + 2:
            try:
                url = (
                    base_url.replace('?', f'/p-{pg}?')
                    if '?' in base_url else f"{base_url}/p-{pg}"
                )
                logger.info(f"Loading page {pg}: {url}")
                page.goto(url, wait_until='domcontentloaded', timeout=45000)
                try:
                    page.wait_for_selector('div.item', timeout=12000)
                except PlaywrightTimeoutError:
                    logger.info(f"No more listings on page {pg}")
                    break
                scraped = parse_page(page.content(), pg_num=pg)
                if scraped == 0:
                    logger.info(f"Empty page {pg} → stopping")
                    break
                time.sleep(2.8 + random.uniform(0, 1.8))
            except Exception as scrape_err:
                logger.error(f"Error on page {pg}: {scrape_err}")
                break
            pg += 1

    except Exception as e:
        logger.critical(f"Critical scraping error: {e}")
    finally:
        browser.close()
        logger.info("Browser closed")

# ================= PROCESSING =================
if not listings:
    logger.warning("No listings scraped → exiting")
    exit()

df_new = pd.DataFrame(listings)
df_sold_now = pd.DataFrame()

# Selenium: price history + images
selenium_results = scrape_site_price_histories_selenium(df_new[COL_LINK].tolist())
df_new[COL_SITE_PRICE_HISTORY] = df_new[COL_LINK].map(
    lambda u: selenium_results.get(u, {}).get("price_history", "")
)
# === DEBUG: проверка дали има данни ===
num_with_history = df_new[COL_SITE_PRICE_HISTORY].str.strip().astype(bool).sum()
logger.info(f"Извлечени ценови истории: {num_with_history} от {len(df_new)} обяви")

df_new[COL_IMAGES] = df_new[COL_LINK].map(
    lambda u: selenium_results.get(u, {}).get("images", "")
)

# Load history
df_history = pd.DataFrame()
if os.path.exists(HISTORY_FILE):
    try:
        df_history = pd.read_parquet(HISTORY_FILE)
        logger.info(f"Loaded history: {len(df_history)} rows")
    except Exception as e:
        logger.error(f"Error reading parquet: {e}")

# Ensure all columns exist in history
if not df_history.empty:
    df_history = deduplicate_history(df_history, COL_LINK, COL_PRICE_HISTORY)
    df_history[COL_PRICE_HISTORY] = df_history[COL_PRICE_HISTORY].fillna("")
    for col, default in [
        (COL_SITE_PRICE_HISTORY, ""),
        (COL_SOLD, False),
        (COL_FIRST_SEEN, ""),
        (COL_IMAGES, ""),
    ]:
        if col not in df_history.columns:
            df_history[col] = default
    df_history[COL_SOLD] = df_history[COL_SOLD].fillna(False)

df_all = df_history.copy()
if not df_all.empty:
    df_all = df_all.set_index(COL_LINK, drop=False)
    for col, default in [
        (COL_SOLD, False),
        (COL_SITE_PRICE_HISTORY, ""),
        (COL_FIRST_SEEN, ""),
        (COL_IMAGES, ""),
    ]:
        if col not in df_all.columns:
            df_all[col] = default

# New listings
df_new_only = (
    df_new[~df_new[COL_LINK].isin(df_all[COL_LINK])]
    if not df_all.empty else df_new.copy()
)

# Changed prices
df_changed = pd.DataFrame()
if not df_history.empty:
    merged = df_new.merge(
        df_history[[COL_LINK, COL_PRICE, COL_LOCATION, COL_SIZE, COL_PRICE_HISTORY]],
        on=COL_LINK,
        how='inner',
        suffixes=('_new', '_old')
    )
    changed_mask = merged[f'{COL_PRICE}_new'] != merged[f'{COL_PRICE}_old']
    changed = merged[changed_mask].copy()

    if not changed.empty:
        def update_history(hist_row):
            hist = hist_row[f'{COL_PRICE_HISTORY}_old']
            old_entry = format_price_history_entry(hist_row[f'{COL_PRICE}_old'], "before")
            if old_entry and old_entry not in hist:
                return f"{hist} → {old_entry}" if hist.strip() else old_entry
            return hist


        changed[f'{COL_PRICE_HISTORY}_updated'] = changed.apply(update_history, axis=1)
        changed[f'{COL_PRICE_PER_SQM}_new'] = (
            round(changed[f'{COL_PRICE}_new'] / changed[f'{COL_SIZE}_new'], 2)
            if f'{COL_SIZE}_new' in changed.columns else None
        )
        df_changed = changed[[
            f'{COL_LOCATION}_old',
            f'{COL_PRICE}_old',
            f'{COL_PRICE}_new',
            f'{COL_SIZE}_new',
            f'{COL_PRICE_PER_SQM}_new',
            COL_LINK,
            COL_SITE_PRICE_HISTORY,
            f'{COL_PRICE_HISTORY}_updated',
        ]].rename(columns={
            f'{COL_LOCATION}_old': COL_LOCATION,
            f'{COL_SIZE}_new': COL_SIZE,
            f'{COL_PRICE_HISTORY}_updated': COL_PRICE_HISTORY,
        })
        logger.info(f"Found {len(df_changed)} price changes")

# Update df_all
for _, row in df_new.iterrows():
    link = row[COL_LINK]
    row_dict = row.to_dict()

    if link in df_all.index:
        old_price = df_all.at[link, COL_PRICE]
        old_scraped_date = df_all.at[link, COL_SCRAPED_DATE]
        new_price = row_dict.get(COL_PRICE)
        new_scraped_date = TODAY

        # Preserve first_seen date
        row_dict[COL_FIRST_SEEN] = df_all.at[link, COL_FIRST_SEEN] or old_scraped_date

        # Preserve images if already downloaded
        existing_images = df_all.at[link, COL_IMAGES] if COL_IMAGES in df_all.columns else ""
        if existing_images and not row_dict.get(COL_IMAGES):
            row_dict[COL_IMAGES] = existing_images

        for col in df_all.columns:
            if col in row_dict and col != COL_PRICE_HISTORY and col != COL_FIRST_SEEN:
                df_all.at[link, col] = row_dict[col]
        if COL_SITE_PRICE_HISTORY in row_dict:
            df_all.at[link, COL_SITE_PRICE_HISTORY] = row_dict[COL_SITE_PRICE_HISTORY]

        if pd.notna(new_price) and pd.notna(old_price) and old_price != new_price:
            current_hist = df_all.at[link, COL_PRICE_HISTORY] or ""
            old_entry = format_price_history_entry(old_price, old_scraped_date)
            if old_entry and old_entry not in current_hist:
                current_hist = f"{current_hist} → {old_entry}" if current_hist.strip() else old_entry
            new_entry = format_price_history_entry(new_price, new_scraped_date)
            if new_entry and new_entry not in current_hist:
                current_hist = f"{current_hist} → {new_entry}" if current_hist.strip() else new_entry
            df_all.at[link, COL_PRICE_HISTORY] = current_hist
    else:
        # new listing
        if pd.notna(row_dict.get(COL_PRICE)):
            row_dict[COL_PRICE_HISTORY] = format_price_history_entry(row_dict[COL_PRICE], TODAY)
        row_dict[COL_SOLD] = False
        row_dict[COL_FIRST_SEEN] = TODAY
        df_all = pd.concat([df_all, pd.DataFrame([row_dict])], ignore_index=True)

# Detect sold
if not df_history.empty:
    base_unsold = df_history[~df_history[COL_SOLD].fillna(False)]
    if not base_unsold.empty:
        sold_links = set(base_unsold[COL_LINK]) - set(df_new[COL_LINK])
        if sold_links:
            df_sold_now = base_unsold[base_unsold[COL_LINK].isin(sold_links)].copy()
            if COL_SOLD not in df_all.columns:
                df_all[COL_SOLD] = False
            df_all.loc[df_all[COL_LINK].isin(sold_links), COL_SOLD] = True

# Final cleanup
df_all = df_all.drop_duplicates(subset=[COL_LINK], keep='last').reset_index(drop=True)

logger.info(
    f"New: {len(df_new_only)}  |  Changed: {len(df_changed)}  |  "
    f"Sold: {len(df_sold_now)}  |  Total unique: {len(df_all)}"
)

# Save data
df_all.to_parquet(HISTORY_FILE, index=False)
df_all.to_csv("all_listings_history.csv", index=False, encoding='utf-8-sig')

# ================= GENERATE HTML DASHBOARD =================
generate_html(df_all, NOW_STR)

# ================= EXCEL EXPORT =================
df_export = df_all.copy()
df_export['Current Price'] = df_export[COL_PRICE].apply(lambda x: f"{int(round(x)):,} €" if pd.notna(x) else "")
df_export['Price per m²'] = df_export[COL_PRICE_PER_SQM].apply(
    lambda x: f"{int(round(x)):,} €/m²" if pd.notna(x) else "")
df_export['Size'] = df_export[COL_SIZE].apply(lambda x: f"{int(x)} m²" if pd.notna(x) else "")
df_export[COL_PRICE_HISTORY] = df_export.apply(
    lambda r: append_current_if_needed(r[COL_PRICE_HISTORY], r[COL_PRICE], r[COL_SCRAPED_DATE]), axis=1
)
df_export = df_export.rename(columns={
    COL_PRICE_HISTORY: 'Price History',
    COL_SITE_PRICE_HISTORY: 'Site price history',
    COL_PRICE: 'Current Price (numeric)',
    COL_PRICE_PER_SQM: 'Price per m² (numeric)',
    COL_SIZE: 'Size (numeric)',
    COL_SCRAPED_DATE: 'Scraped Date',
    COL_FIRST_SEEN: 'First Seen Date',
    COL_LOCATION: 'Location',
    COL_TITLE: 'Title',
    COL_FLOOR: 'Floor',
    COL_TOTAL_FLOORS: 'Total Floors',
    COL_YEAR: 'Year Built',
    COL_IMAGES: 'Image Paths',
})
df_export.to_excel(excel_file, index=False, engine='openpyxl')

try:
    wb = load_workbook(excel_file)
    ws = wb.active
    sold_col_idx = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col_idx).value == COL_SOLD:
            sold_col_idx = col_idx
            break
    if sold_col_idx:
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=sold_col_idx).value:
                for col_idx in range(1, ws.max_column + 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.font = Font(strikethrough=True, color="FF777777")
    wb.save(excel_file)
except Exception as e:
    logger.error(f"Excel formatting error: {e}")

logger.info(f"Excel saved: {excel_file}")

# ================= EMAIL =================
if len(df_new_only) > 0 or len(df_changed) > 0 or len(df_sold_now) > 0:

    def fmt_p(x):
        return f"{x:,.0f} €" if pd.notna(x) else "—"


    def fmt_s(x):
        return f"{int(x)} m²" if pd.notna(x) else "—"


    def fmt_m(x):
        return f"{int(round(x))} €/m²" if pd.notna(x) else "—"


    CSS = """<style>
body{font-family:Arial,sans-serif;line-height:1.6;color:#333;background:#f9f9f9}
.wrap{max-width:900px;margin:0 auto;background:#fff;border-radius:8px;overflow:hidden;border:1px solid #e0e0e0}
.hdr{background:#1a1d27;color:#fff;padding:20px 28px}
.hdr h1{font-size:18px;margin:0}
.hdr p{font-size:12px;color:#7a7d9a;margin:4px 0 0}
.body{padding:24px 28px}
h3{color:#444;font-size:14px;margin:24px 0 8px;border-bottom:2px solid #f0f0f0;padding-bottom:6px}
.imot-table{border-collapse:collapse;width:100%;font-size:13px;margin-bottom:8px}
.imot-table th{background:#f5f5f5;font-size:11px;text-transform:uppercase;letter-spacing:.04em;color:#666;padding:8px 12px;text-align:left;border-bottom:2px solid #e0e0e0}
.imot-table td{padding:9px 12px;border-bottom:1px solid #f0f0f0;vertical-align:top}
.imot-table tr:hover td{background:#fafafa}
a{color:#4f9cf9;text-decoration:none}
.pill{display:inline-block;padding:2px 10px;border-radius:99px;font-size:12px;font-weight:600}
.pill.new{background:#dcfce7;color:#166534}
.pill.chg{background:#fef3c7;color:#92400e}
.pill.sld{background:#fee2e2;color:#991b1b}
.sold-table td{color:#999;text-decoration:line-through}
.ftr{background:#f5f5f5;padding:12px 28px;font-size:12px;color:#999;text-align:center}
</style>"""


    def to_html_table(df_in, cols, headers, extra_class=""):
        tbl_head = "<tr>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr>"
        tbl_rows = []
        for _, email_row in df_in.iterrows():
            cells = []
            for col_name in cols:
                v = email_row.get(col_name, "")
                if col_name == COL_LINK:
                    cells.append(f"<td><a href='{v}' target='_blank'>Виж →</a></td>")
                elif col_name == COL_PRICE:
                    cells.append(f"<td>{fmt_p(v)}</td>")
                elif col_name == COL_SIZE:
                    cells.append(f"<td>{fmt_s(v)}</td>")
                elif col_name == COL_PRICE_PER_SQM:
                    cells.append(f"<td>{fmt_m(v)}</td>")
                elif col_name == 'Price_EUR_old':
                    cells.append(f"<td style='text-decoration:line-through;color:#999'>{fmt_p(v)}</td>")
                else:
                    cells.append(f"<td>{v if pd.notna(v) and v != '' else '—'}</td>")
            tbl_rows.append("<tr>" + "".join(cells) + "</tr>")
        cls = f"imot-table {extra_class}".strip()
        return f"<table class='{cls}'><thead>{tbl_head}</thead><tbody>{''.join(tbl_rows)}</tbody></table>"


    new_section = changed_section = sold_section = ""

    if not df_new_only.empty:
        tbl = to_html_table(df_new_only,
                            [COL_LOCATION, COL_PRICE, COL_SIZE, COL_PRICE_PER_SQM, COL_SITE_PRICE_HISTORY, COL_LINK],
                            ["Локация", "Цена", "Площ", "€/m²", "История (сайт)", ""])
        new_section = f"""<h3><span class="pill new">НОВИ</span> &nbsp;{len(df_new_only)} обяви &mdash; {TODAY}</h3>{tbl}"""

    if not df_changed.empty:
        df_ch = df_changed.copy()
        if COL_PRICE_PER_SQM not in df_ch.columns and f'{COL_PRICE_PER_SQM}_new' in df_ch.columns:
            df_ch = df_ch.rename(columns={f'{COL_PRICE_PER_SQM}_new': COL_PRICE_PER_SQM})
        if COL_PRICE in df_ch.columns and COL_SIZE in df_ch.columns and COL_PRICE_PER_SQM not in df_ch.columns:
            df_ch[COL_PRICE_PER_SQM] = (df_ch[COL_PRICE] / df_ch[COL_SIZE]).round(2)
        tbl = to_html_table(df_ch,
                            [COL_LOCATION, COL_PRICE, COL_SIZE, COL_PRICE_PER_SQM, COL_PRICE_HISTORY, COL_SITE_PRICE_HISTORY, COL_LINK],
                            ["Локация", "Нова цена", "Площ", "€/m²", "История на цената", "История (сайт)", ""])
        changed_section = f"""<h3><span class="pill chg">ПРОМЯНА В ЦЕНА</span> &nbsp;{len(df_changed)} обяви &mdash; {TODAY}</h3>{tbl}"""

    if not df_sold_now.empty:
        tbl = to_html_table(df_sold_now,
                            [COL_LOCATION, COL_PRICE, COL_SIZE, COL_PRICE_PER_SQM, COL_SITE_PRICE_HISTORY, COL_LINK],
                            ["Локация", "Последна цена", "Площ", "€/m²", "История (сайт)", ""],
                            extra_class="sold-table")
        sold_section = f"""<h3><span class="pill sld">ПРОДАДЕНИ</span> &nbsp;{len(df_sold_now)} обяви &mdash; {TODAY}</h3>{tbl}"""

    subject_parts = []
    if not df_new_only.empty:   subject_parts.append(f"{len(df_new_only)} НОВИ")
    if not df_changed.empty:    subject_parts.append(f"{len(df_changed)} ПРОМЯНА")
    if not df_sold_now.empty:   subject_parts.append(f"{len(df_sold_now)} ПРОДАДЕНИ")
    subject = f"Имоти – {' · '.join(subject_parts)} – {TODAY}"

    email_html = f"""<html><head><meta charset="utf-8">{CSS}</head><body>
<div class="wrap">
  <div class="hdr">
    <h1>🏠 Имоти · сем. Кирилови</h1>
    <p>Автоматичен отчет · {NOW_STR}</p>
  </div>
  <div class="body">
    {new_section}{changed_section}{sold_section}
    <p style="margin-top:24px;font-size:13px;color:#666">
      Общо уникални обяви в базата: <strong>{len(df_all)}</strong><br>
      Пълният списък е прикачен като Excel.
    </p>
  </div>
  <div class="ftr">Тук може да е вашият следващ ДОМ! 🏡</div>
</div>
</body></html>"""

    msg = MIMEMultipart()
    msg['From'] = SENDER_EMAIL
    msg['To'] = ", ".join(RECEIVERS)
    msg['Subject'] = subject
    msg.attach(MIMEText(email_html, "html", _charset="utf-8"))

    try:
        with open(excel_file, 'rb') as f:
            part = MIMEApplication(f.read(), Name=excel_file)
            part['Content-Disposition'] = f'attachment; filename="{excel_file}"'
            msg.attach(part)
    except Exception as attach_err:
        logger.error(f"Failed to attach Excel: {attach_err}")

    try:
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.send_message(msg)
        server.quit()
        logger.info("✔ Email sent")
    except Exception as smtp_err:
        logger.error(f"❌ Email error: {smtp_err}")
else:
    logger.info("No changes → email not sent")

logger.info("=== Script finished ===")